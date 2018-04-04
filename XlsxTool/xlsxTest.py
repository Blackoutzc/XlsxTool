# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import copy
import sys
from collections import Iterable
from openpyxl.styles import Font, PatternFill
import xlrd
#RGB颜色对照表 http://www.114la.com/other/rgb.htm


class Encoding(object):
    @staticmethod
    def to_unicode(string):
        if isinstance(string, str):
            return string.decode("utf-8")
        else:
            return string

    @staticmethod
    def to_str(string):
        if isinstance(string, str):
            return string
        else:
            return string.encode("utf-8")


class AbstractSheet(object):
    def __init__(self, worksheet):
        assert worksheet is not None and isinstance(worksheet, openpyxl.worksheet.worksheet.Worksheet)
        self._sheet = worksheet
        self.sheet = None

    @property
    def max_row(self):
        """Return the last row number which has contents"""
        return self._sheet.max_row

    @property
    def max_column(self):
        """Return the last column number which has contents"""
        return self._sheet.max_column

    @property
    def raw_sheet(self):
        """Return the openpyxl sheet , which is an instance of openpyxl.worksheet.worksheet.Worksheet"""
        return self._sheet

    def value(self, row, column):
        """Returns the value specified by row number and column number"""
        cell_value = self._sheet.cell(row=row, column=column).value
        if cell_value is not None:
            if isinstance(cell_value, (str, unicode)):
                return cell_value.strip()
            else:
                return cell_value
        else:
            return None

    def get_cell(self, row, column):
        """Return an openpyxl.cell.cell.Cell object"""
        return self._sheet.cell(row=row, column=column)

    @staticmethod
    def __construct_style(style=None):
        style_copy = copy.deepcopy(style)
        fill, font = None, None
        if style_copy is not None and isinstance(style_copy, dict):
            if "background" in style_copy:
                fill = PatternFill("solid", fgColor=style_copy["background"])
                style_copy.pop("background")
            font = Font(**style_copy)
        return font, fill

    @staticmethod
    def __set_styles(cell_inst, font, fill):
        if font is not None:
            cell_inst.font = font
        if fill is not None:
            cell_inst.fill = fill

    def set_style_with_cell(self, cell_inst, style=None):
        font, fill = self.__construct_style(style)
        self.__set_styles(cell_inst, font, fill)

    def set_style(self, row, column, style=None):
        """Set the specified cell's style, style should be a dictionary as
        {"color": colors.RED, "background": "7CCD7C", "bold": True}
        We only support those styles which is frequently used, if it is beyond
        your needs please use the method supplied by openpyxl, we can easily
        access openpyxl.cell.cell.Cell object via get_cell(row, value)
        """
        cell_inst = self.get_cell(row, column)
        self.set_style_with_cell(cell_inst, style)

    def get_cells(self, first_row, row_length, first_column, column_length):
        return [self.get_cell(row, column) for row in xrange(first_row, first_row+row_length) for column in xrange(first_column, first_column+column_length)]

    def set_range_style(self, first_row, row_length, first_column, column_length, style=None):
        font, fill = self.__construct_style(style)
        cells_collection = self.get_cells(first_row, row_length, first_column, column_length)
        for cell in cells_collection:
            self.__set_styles(cell, font, fill)

    def append(self, values):
        """Append values after the last row which has contents
            :param values: values to be added
            :type values: list or tuple
        """
        assert isinstance(values, (list, tuple))
        self._sheet.append(values)

    def set_value(self, row, column, value, style=None):
        """Write value into cell specified by row(int) and column(int)"""
        cell_inst = self.get_cell(row, column)
        font, fill = self.__construct_style(style)
        self.__set_value_s(cell_inst, value, font, fill)

    def __set_value_s(self, cell_inst, value, font, fill):
        cell_inst.value = value
        self.__set_styles(cell_inst, font, fill)

    def set_row_values(self, values, row_num, start_column_index=1, style=None):
        """Set a whole row's values
            :param values: values to be added
            :type values: list or tuple
        """
        if isinstance(values, (list, tuple)):
            column_num = start_column_index
            font, fill = self.__construct_style(style)
            for value in values:
                assert not isinstance(value, Iterable) or isinstance(value, (str, unicode))

                cell = self.get_cell(row_num, column_num)
                self.__set_value_s(cell, value, font, fill)
                column_num += 1

    def append_column(self, values, style=None):
        self.set_column_values(values, self.max_column, style=style)

    def set_column_values(self, values, column_num, start_row_index=1, style=None):
        """Set a whole columns's values"""
        if isinstance(values, (list, tuple)):
            row_num = start_row_index
            font, fill = self.__construct_style(style)
            for value in values:
                # print value
                assert not isinstance(value, Iterable) or isinstance(value, (str, unicode))
                cell = self.get_cell(row_num, column_num)
                self.__set_value_s(cell, value, font, fill)
                row_num += 1

    def column_values(self, column, start_row_num=1):
        value_collection = []
        for i in xrange(start_row_num, self.max_row + 1):
            value_collection.append(self.value(i, column))
        return self.__ditch_end_none(value_collection)

    def row_values(self, row, start_column_num=1):
        value_collection = []
        for i in xrange(start_column_num, self.max_column + 1):
            value_collection.append(self.value(row, i))
        return self.__ditch_end_none(value_collection)

    @staticmethod
    def __ditch_end_none(list_value):
        for i in xrange(len(list_value) - 1, 0, -1):
            if list_value[i] is None or list_value[i] == "":
                list_value.pop()
            else:
                break
        return list_value


class XlsxHandler(AbstractSheet):
    def __init__(self, open_file_name=None, sheetname=None):
        self._open_file_name = Encoding.to_unicode(open_file_name)
        self._current_sheet_name = Encoding.to_unicode(sheetname)
        self._workbook = None
        self._current_sheet = None
        self._sheets_name = None
        if self._open_file_name is not None:
            if self._open_file_name.endswith(".xls"):
                self._convert_xls_to_xlsx()
            else:
                self._open_xlsxfile()
        else:
            self._create()
        self._current_sheet = self.worksheet(self._current_sheet_name)
        self._worksheet_is_valid(self._current_sheet)
        super(XlsxHandler, self).__init__(self._current_sheet)

    def __enter__(self):
        """in order to support context manager"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    @property
    def file_name(self):
        """Return the file path"""
        return self._open_file_name

    def _convert_xls_to_xlsx(self):
        """convert .xls file into .xlsx file"""
        xls_book = xlrd.open_workbook(filename=self._open_file_name)
        self._create()
        for i in xrange(0, xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(i)
            sheet = self.worksheet() if i == 0 else self._workbook.create_sheet()
            sheet.title = xls_sheet.name

            for row in xrange(0, xls_sheet.nrows):
                for col in xrange(0, xls_sheet.ncols):
                    sheet.cell(row=row+1, column=col+1).value = xls_sheet.cell_value(row, col) \
                        if xls_sheet.cell_value(row, col) != '' else None
        self._sheets_name = self._workbook.sheetnames

    @property
    def sheet_name(self):
        pass

    def _create(self):
        self._workbook = Workbook()

    @staticmethod
    def is_name_valid(file_path):
        """To tell if the given file_path is valid for handling"""
        try:
            assert file_path and isinstance(file_path, (str, unicode)) and file_path.endswith(".xlsx")
            return True
        except AssertionError as e:
            print "oops, it seems like {} is not the right excel name".format(file_path)
            return False

    def _open_xlsxfile(self):
        self._workbook = load_workbook(self._open_file_name)
        self._sheets_name = self._workbook.sheetnames

    def _sheetname_is_valid(self, sheetname):
        try:
            assert Encoding.to_unicode(sheetname) in self._sheets_name
        except AssertionError:
            #print "Error: sheetname {0} is not valid".format(sheetname)
            sys.exit("Error: sheetname {0} is not valid".format(sheetname))

    @staticmethod
    def _worksheet_is_valid(worksheet):
        try:
            assert isinstance(worksheet, openpyxl.worksheet.worksheet.Worksheet)
        except AssertionError:
            sys.exit("Error: worksheet is not available")

    @staticmethod
    def _workbook_is_valid(workbook):
        try:
            assert isinstance(workbook, openpyxl.workbook.workbook.Workbook)
        except AssertionError:
            sys.exit("Error: workbook is not available")

    @property
    def workbook(self):
        """Return the instance of openpyxl.workbook.workbook.Workbook"""
        return self._workbook

    def worksheet(self, sheetname=None):
        """Return the instance of openpyxl.worksheet.worksheet.Worksheet"""
        self._workbook_is_valid(self._workbook)
        return self._workbook.active if sheetname is None else self._workbook[sheetname]

    def sheet(self, sheetname=None):
        worksheet = self.worksheet(Encoding.to_unicode(sheetname))
        self._worksheet_is_valid(worksheet)
        return AbstractSheet(worksheet)

    def change_sheet(self, sheetname=None):
        self._sheetname_is_valid(sheetname)
        self._current_sheet = self.worksheet(Encoding.to_unicode(sheetname))
        self._sheet = self._current_sheet

    def create_sheet(self, sheetname=None):
        sheetname = Encoding.to_unicode(sheetname)
        self._workbook_is_valid(self._workbook)
        self._workbook.create_sheet(title=sheetname)
        self._sheets_name = self._workbook.sheetnames
        return AbstractSheet(self._workbook[sheetname])

    def sheet_names(self):
        return self._sheets_name

    def save(self, filename):
        """we only support saving into .xlsx file"""
        try:
            assert Encoding.to_unicode(filename).endswith(".xlsx")
        except AssertionError:
            sys.exit("Error : Saving File Name is Wrong")
        self._workbook.save(Encoding.to_unicode(filename))

    def run(self):
        self.process()

    def process(self):
        pass